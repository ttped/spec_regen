import openpyxl
import docx
from docx_writer import add_excel_table_to_docx
from docx_writer import add_isolated_landscape_table, add_excel_table_to_docx

def read_excel_with_metadata(file_path: str) -> dict:
    """
    Reads an Excel file and extracts text values and column width metadata.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    table_data = {"columns": [], "rows": []}

    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        width = ws.column_dimensions[col_letter].width
        
        if width is None:
            width = 8.43
            
        table_data["columns"].append({"width": width})

    for row in ws.iter_rows(values_only=True):
        table_data["rows"].append(list(row))

    return table_data

def export_excel_to_word(excel_path: str, word_output_path: str):
    """
    Reads Excel data and writes it to Word, applying landscape for wide tables.
    """
    table_data = read_excel_with_metadata(excel_path)
    
    doc = docx.Document()
    
    num_columns = len(table_data.get("columns", []))
    
    # Route to landscape if the table is exceptionally wide
    if num_columns > 7:
        add_isolated_landscape_table(doc, table_data)
    else:
        add_excel_table_to_docx(doc, table_data)
        
    doc.save(word_output_path)
    print(f"Successfully converted {excel_path} to {word_output_path}")

if __name__ == "__main__":
    input_file = "sample_input.xlsx"
    output_file = "excel_output.docx"
    
    export_excel_to_word(input_file, output_file)
