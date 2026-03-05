"""
table_iris_processor.py - Process tables using IRIS table extraction deliverables.

Replaces the LLM-based table_processor_agent. For each table element in the
document stream, checks whether a pre-built Excel file exists in the IRIS
deliverable. If found, reads it directly to populate table_data on the element.
If not found, falls back to the IRIS metadata JSON for direct conversion.
If neither exists, the element keeps its image for fallback rendering in docx_writer.

Expected deliverable layout:
    table_jsons_dir/
        <doc_name>/
            excel/
                doc_name_page045_tab_layout_0.xlsx
                doc_name_page045_tab_caption_1.xlsx
            table_jsons/
                doc_name_page045_tab_layout_0.json   (metadata fallback)
                doc_name_page045_tab_caption_1.json

YOLO crops are named:  {doc}_tab_p{page}_{id}.jpg
IRIS files are named:  {doc}_page{page}_tab_layout_{id}  (or tab_caption_{id})

Both are decomposed into a canonical key (doc, page, table_id) for matching.
"""

import os
import json
import re
from pathlib import Path
from typing import List, Dict, Optional


# ---------------------------------------------------------------------------
# Canonical key extraction — bridges naming conventions
# ---------------------------------------------------------------------------

# YOLO crop names: {doc}_tab_p{page}_{id}
_YOLO_PATTERN = re.compile(
    r'^(.+?)_tab_p(\d+)_(\d+)$',
    re.IGNORECASE,
)

# IRIS names: {doc}_page{page}_tab_layout_{id} or {doc}_page{page}_tab_caption_{id}
_IRIS_PATTERN = re.compile(
    r'^(.+?)_page_?(\d+)_?tab_(?:layout|caption)_(\d+)$',
    re.IGNORECASE,
)

# Legacy IRIS pattern: {doc}_page{page}_tab_{id}
_IRIS_LEGACY_PATTERN = re.compile(
    r'^(.+?)_page_?(\d+)_?tab_(\d+)$',
    re.IGNORECASE,
)


def _normalize_stem(name: str) -> str:
    """Collapse spaces, underscores, hyphens and case for fuzzy matching."""
    return name.replace(" ", "_").replace("-", "_").lower()


def _extract_canonical_key(stem: str) -> Optional[tuple]:
    """
    Extract (normalized_doc, page_int, table_id_int) from any naming convention.
    Returns None if the stem doesn't match a recognizable pattern.
    """
    normalized = _normalize_stem(stem)

    # Extract page number: look for p or page followed by optional underscore and digits
    page_match = re.search(r'(?:^|_)(?:p|page)_?(\d+)', normalized)
    if not page_match:
        return None
    page = int(page_match.group(1))

    # Extract table ID: usually the last sequence of digits in the string
    id_match = re.search(r'_(\d+)$', normalized)
    if not id_match:
        return None
    table_id = int(id_match.group(1))

    # Extract doc name: everything before the table/page markers
    doc_match = re.match(r'^(.*?)(?:_tab|_table|(?:^|_)(?:p|page)_?\d+)', normalized)
    doc = doc_match.group(1) if doc_match and doc_match.group(1) else "unknown_doc"
    doc = doc.rstrip("_")

    return (doc, page, table_id)



# ---------------------------------------------------------------------------
# Index builders — scan deliverable for Excel and JSON files
# ---------------------------------------------------------------------------

def _build_excel_index(table_jsons_dir: str) -> Dict[str, dict]:
    """
    Scan table_jsons_dir for Excel files. Handles multiple layouts.
    Builds canonical, page-based, doc-agnostic page-based, and stem indexes.
    """
    canonical_index = {}
    page_index = {}
    agnostic_page_index = {}
    stem_index = {}
    file_count = 0

    if not os.path.isdir(table_jsons_dir):
        print(f"    [Warning] table_jsons_dir not found: {table_jsons_dir}")
        return {"_canonical": {}, "_page": {}, "_agnostic_page": {}, "_stem": {}}

    def _index_file(filepath, stem):
        nonlocal file_count
        file_count += 1

        key = _extract_canonical_key(stem)
        if key:
            canonical_index[key] = filepath
            doc, page, table_id = key
            
            page_key = (doc, page)
            if page_key not in page_index:
                page_index[page_key] = []
            page_index[page_key].append((table_id, filepath))
            
            if page not in agnostic_page_index:
                agnostic_page_index[page] = []
            agnostic_page_index[page].append((table_id, filepath))

        stem_index[_normalize_stem(stem)] = filepath

    # Structured scan: */excel/*.xlsx
    for entry in os.listdir(table_jsons_dir):
        excel_dir = os.path.join(table_jsons_dir, entry, "excel")
        if not os.path.isdir(excel_dir):
            continue

        for filename in os.listdir(excel_dir):
            if not filename.endswith(".xlsx") or filename.startswith("~$"):
                continue
            _index_file(os.path.join(excel_dir, filename), os.path.splitext(filename)[0])

    # Fallback: recursive walk for any .xlsx if structured scan found nothing
    if file_count == 0:
        for root, _dirs, files in os.walk(table_jsons_dir):
            for filename in files:
                if not filename.endswith(".xlsx") or filename.startswith("~$"):
                    continue
                _index_file(os.path.join(root, filename), os.path.splitext(filename)[0])

    # Sort page_index and agnostic_page_index entries by table_id for positional matching
    for page_key in page_index:
        page_index[page_key].sort(key=lambda x: x[0])
    for page_num in agnostic_page_index:
        agnostic_page_index[page_num].sort(key=lambda x: x[0])

    print(f"  - Scanned Excel files: {file_count} in {table_jsons_dir}")

    return {
        "_canonical": canonical_index, 
        "_page": page_index, 
        "_agnostic_page": agnostic_page_index,
        "_stem": stem_index
    }


def _build_json_index(table_jsons_dir: str) -> Dict[str, dict]:
    """
    Scan table_jsons_dir for IRIS metadata JSON files (fallback).
    Builds canonical, page-based, doc-agnostic page-based, and stem indexes.
    """
    canonical_index = {}
    page_index = {}
    agnostic_page_index = {}
    stem_index = {}
    file_count = 0

    if not os.path.isdir(table_jsons_dir):
        return {"_canonical": {}, "_page": {}, "_agnostic_page": {}, "_stem": {}}

    def _index_file(filepath, stem):
        nonlocal file_count
        file_count += 1

        key = _extract_canonical_key(stem)
        if key:
            canonical_index[key] = filepath
            doc, page, table_id = key
            
            page_key = (doc, page)
            if page_key not in page_index:
                page_index[page_key] = []
            page_index[page_key].append((table_id, filepath))
            
            if page not in agnostic_page_index:
                agnostic_page_index[page] = []
            agnostic_page_index[page].append((table_id, filepath))

        stem_index[_normalize_stem(stem)] = filepath

    # Structured scan: */table_jsons/*.json
    for entry in os.listdir(table_jsons_dir):
        json_dir = os.path.join(table_jsons_dir, entry, "table_jsons")
        if not os.path.isdir(json_dir):
            continue

        for filename in os.listdir(json_dir):
            if not filename.endswith(".json"):
                continue
            _index_file(os.path.join(json_dir, filename), os.path.splitext(filename)[0])

    # Fallback: recursive walk for any .json
    if file_count == 0:
        for root, _dirs, files in os.walk(table_jsons_dir):
            for filename in files:
                if not filename.endswith(".json"):
                    continue
                _index_file(os.path.join(root, filename), os.path.splitext(filename)[0])

    # Sort page_index and agnostic_page_index entries by table_id
    for page_key in page_index:
        page_index[page_key].sort(key=lambda x: x[0])
    for page_num in agnostic_page_index:
        agnostic_page_index[page_num].sort(key=lambda x: x[0])

    return {
        "_canonical": canonical_index, 
        "_page": page_index, 
        "_agnostic_page": agnostic_page_index,
        "_stem": stem_index
    }



def _lookup(stem: str, index: dict, position_on_page: int = 0) -> Optional[str]:
    """
    Look up a file path for a YOLO crop stem.
    
    Matching priority:
    1. Exact canonical key (doc, page, id)
    2. Page-based positional match (doc, page) + nth table on that page
    3. Doc-agnostic page-based positional match (page) + nth table
    4. Normalized stem — direct name match fallback
    """
    canonical_index = index.get("_canonical", {})
    page_index = index.get("_page", {})
    agnostic_page_index = index.get("_agnostic_page", {})
    stem_index = index.get("_stem", {})

    key = _extract_canonical_key(stem)
    
    if key:
        doc, page, _yolo_id = key

        # 1. Exact canonical key match
        if key in canonical_index:
            return canonical_index[key]

        # 2. Page-based positional match (includes doc)
        page_key = (doc, page)
        if page_key in page_index:
            entries = page_index[page_key]  # sorted by IRIS table_id
            if position_on_page < len(entries):
                return entries[position_on_page][1]
            elif len(entries) == 1:
                return entries[0][1]

        # 3. Doc-agnostic page-based match (fixes cases where YOLO doc != IRIS doc)
        if page in agnostic_page_index:
            entries = agnostic_page_index[page]
            if position_on_page < len(entries):
                return entries[position_on_page][1]
            elif len(entries) == 1:
                return entries[0][1]

    # 4. Fallback: direct normalized stem match
    normalized = _normalize_stem(stem)
    return stem_index.get(normalized)


# ---------------------------------------------------------------------------
# Excel → table_data (read pre-built Excel directly)
# ---------------------------------------------------------------------------

def read_excel_to_table_data(excel_path: str) -> Optional[Dict]:
    """
    Read an Excel file and convert to the table_data format used by
    add_excel_table_to_docx / add_isolated_landscape_table.

    Returns:
        {"columns": [{"width": float}, ...], "rows": [[val, ...], ...]}
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    table_data = {"columns": [], "rows": []}

    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        width = ws.column_dimensions[col_letter].width
        if width is None:
            width = 8.43
        table_data["columns"].append({"width": width})

    for row in ws.iter_rows(values_only=True):
        table_data["rows"].append(list(row))

    if not table_data["rows"]:
        return None

    table_data["_num_columns"] = len(table_data["columns"])
    return table_data


# ---------------------------------------------------------------------------
# IRIS JSON → complex_table_schema (fallback if no Excel available)
# ---------------------------------------------------------------------------

def read_iris_json_to_table_data(iris_json_path: str) -> Optional[Dict]:
    """
    Convert an IRIS table extraction JSON directly into the rich format
    expected by complex_table_schema. Fallback when no pre-built Excel exists.
    """
    with open(iris_json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    n_rows = data.get("n_rows", 0)
    n_cols = data.get("n_cols", 0)
    iris_rows = data.get("rows", [])

    if n_rows == 0 or n_cols == 0 or not iris_rows:
        return None

    header_row_indices = set(data.get("header_rows", []))
    context = data.get("context", {})
    context_header_count = context.get("header_rows", 0)

    columns = [{"name": ""} for _ in range(n_cols)]
    covered = [[False] * n_cols for _ in range(n_rows)]
    iris_rows_sorted = sorted(iris_rows, key=lambda r: r.get("row_index", 0))

    output_rows = []
    for iris_row in iris_rows_sorted:
        row_idx = iris_row.get("row_index", 0)
        
        # Safeguard against out-of-bounds rows
        if row_idx >= n_rows or row_idx < 0:
            continue
            
        is_header = row_idx in header_row_indices
        row_cells = [None] * n_cols

        for cell_data in iris_row.get("cells", []):
            col = cell_data.get("col", 0)
            row_span = cell_data.get("row_span", 1)
            col_span = cell_data.get("col_span", 1)
            
            # Safeguard against out-of-bounds columns
            if col < 0 or col >= n_cols:
                continue
                
            text = cell_data.get("text", "")
            alignment = cell_data.get("alignment", "left")

            halign = alignment if alignment in ("left", "center", "right") else "left"
            cell = {"text": str(text) if text is not None else "", "halign": halign}

            if is_header:
                cell["bold"] = True

            # Calculate safe spans (don't exceed table boundaries)
            safe_row_span = min(row_span, n_rows - row_idx)
            safe_col_span = min(col_span, n_cols - col)
            
            # Verify no overlaps in the target merged area
            can_merge = True
            for dr in range(safe_row_span):
                for dc in range(safe_col_span):
                    if covered[row_idx + dr][col + dc]:
                        can_merge = False
                        break
                if not can_merge:
                    break
                    
            if can_merge:
                if safe_col_span > 1:
                    cell["colspan"] = safe_col_span
                if safe_row_span > 1:
                    cell["rowspan"] = safe_row_span
                    
                # Mark area as covered
                for dr in range(safe_row_span):
                    for dc in range(safe_col_span):
                        covered[row_idx + dr][col + dc] = True
            else:
                # Fallback to 1x1 if overlapping to prevent docx merge crashes
                covered[row_idx][col] = True

            row_cells[col] = cell

        # Fill any missing cells in the row with empty strings to guarantee a perfect grid
        for c in range(n_cols):
            if row_cells[c] is None:
                row_cells[c] = {"text": "", "halign": "left"}

        output_row = {"cells": row_cells}
        if is_header:
            output_row["is_header"] = True
        output_rows.append(output_row)

    num_header_rows = max(context_header_count, len(header_row_indices))

    return {
        "columns": columns,
        "rows": output_rows,
        "header_rows": num_header_rows,
        "_num_columns": n_cols,
        "_iris_confidence": data.get("confidence"),
        "_iris_structure_confidence": data.get("structure_confidence"),
        "_iris_table_title": data.get("table_title"),
    }


def _extract_caption_from_iris_json(json_path: str) -> str:
    """
    Extract a table caption/title from an IRIS metadata JSON.
    
    Checks (in priority order):
        1. "table_title" field
        2. "context" -> "title" -> "title" field
        3. "context" -> "caption" field
    
    Returns empty string if no caption found.
    """
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Direct table_title
    title = data.get("table_title", "")
    if title and str(title).strip():
        return str(title).strip()

    # Context block
    context = data.get("context", {})
    if context:
        # context.title.title
        title_block = context.get("title", {})
        if isinstance(title_block, dict):
            title = title_block.get("title", "")
            if title and str(title).strip():
                return str(title).strip()

        # context.caption
        caption = context.get("caption", "")
        if caption and str(caption).strip():
            return str(caption).strip()

    return ""


# ---------------------------------------------------------------------------
# Per-element processing
# ---------------------------------------------------------------------------

def process_table_element(
    element: Dict,
    excel_index: Dict,
    json_index: Dict,
    position_on_page: int = 0,
) -> Dict:
    """
    Process a single table element: find pre-built Excel or fall back to JSON.
    Also extracts caption from IRIS metadata if available.

    Args:
        position_on_page: 0-based index of this table among tables on the same page.
            Used for positional matching when YOLO and IRIS use different table IDs.
    
    Modifies the element in place and returns it.
    """
    export_data = element.get("export") or {}
    image_file = export_data.get("image_file", "")

    if not image_file:
        return element

    image_stem = os.path.splitext(image_file)[0]

    # Try to extract caption from IRIS JSON metadata (regardless of Excel/JSON path)
    json_path = _lookup(image_stem, json_index, position_on_page)
    if json_path and not element.get("caption_text"):
        caption = _extract_caption_from_iris_json(json_path)
        if caption:
            element["caption_text"] = caption
            print(f"    [IRIS] Caption extracted: {caption}")

    # Primary: look for pre-built Excel
    excel_path = _lookup(image_stem, excel_index, position_on_page)
    if excel_path:
        print(f"    [IRIS] Found Excel for: {image_stem}")
        table_data = read_excel_to_table_data(excel_path)
        if table_data:
            num_cols = table_data["_num_columns"]
            num_rows = len(table_data["rows"])
            print(f"    [IRIS] Excel loaded: {num_cols} columns, {num_rows} rows")

            element["table_data"] = table_data
            element["_table_source"] = "iris_excel"
            if num_cols > 7:
                element["_render_landscape"] = True
            return element

    # Fallback: IRIS metadata JSON → complex_table_schema
    if json_path:
        print(f"    [IRIS] No Excel, trying JSON for: {image_stem}")
        table_data = read_iris_json_to_table_data(json_path)
        if table_data:
            num_cols = table_data.get("_num_columns", 0)
            num_rows = len(table_data.get("rows", []))
            print(f"    [IRIS] JSON converted: {num_cols} columns, {num_rows} rows")

            element["table_data"] = table_data
            element["_table_source"] = "iris_json"
            if num_cols > 7:
                element["_render_landscape"] = True
            return element

    return element


# ---------------------------------------------------------------------------
# Main entry point (called from simple_pipeline.py step 6)
# ---------------------------------------------------------------------------

def run_iris_table_processing(
    input_path: str,
    output_path: str,
    table_jsons_dir: str,
    doc_stem: str,
):
    """
    Process all table elements using IRIS deliverables (pre-built Excel preferred).

    Args:
        input_path: Path to the _with_assets.json file
        output_path: Path to write the _with_tables.json file
        table_jsons_dir: Root directory of IRIS deliverable
        doc_stem: Document stem (for logging)
    """
    print(f"  - Reading elements: {input_path}")

    if not os.path.exists(input_path):
        print(f"  - [Error] Input file not found: {input_path}")
        return

    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, dict) and 'elements' in data:
        elements = data.get('elements', [])
        page_metadata = data.get('page_metadata', {})
    else:
        elements = data if isinstance(data, list) else []
        page_metadata = {}

    # Build indexes once
    excel_index = _build_excel_index(table_jsons_dir)
    json_index = _build_json_index(table_jsons_dir)

    excel_count = len(excel_index.get("_canonical", {}))
    json_count = len(json_index.get("_canonical", {}))
    print(f"  - IRIS index: {excel_count} Excel files, {json_count} JSON files")

    # Process each table element, tracking position per page for positional matching
    tables_found = 0
    tables_matched = 0
    page_table_counts = {}  # page_number → count of tables seen so far on that page

    for element in elements:
        if element.get("type") not in ("table", "table_layout"):
            continue

        tables_found += 1
        had_data_before = element.get("table_data") is not None

        # Track position of this table on its page
        page_num = element.get("page_number", 0)
        position_on_page = page_table_counts.get(page_num, 0)
        page_table_counts[page_num] = position_on_page + 1

        process_table_element(element, excel_index, json_index, position_on_page)

        if not had_data_before and element.get("table_data") is not None:
            tables_matched += 1

    print(f"  - Tables in stream: {tables_found}")
    print(f"  - Tables matched: {tables_matched}")
    print(f"  - Tables falling back to image: {tables_found - tables_matched}")

    # Save output
    output_data = {
        "page_metadata": page_metadata,
        "elements": elements,
    }

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=4)

    print(f"  - Saved to: {output_path}")